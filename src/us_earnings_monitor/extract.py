from __future__ import annotations

import io
import os
import zipfile
from pathlib import PurePosixPath
from xml.etree import ElementTree as ET

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from pypdf import PdfReader

from .document_compressor import compress_pdf_pages, compress_text
from .models import Disclosure, Evidence

_KEYWORDS = (
    "revenue", "net sales", "operating income", "net income", "eps", "guidance",
    "outlook", "orders", "demand", "margin", "cash flow", "capital expenditure", "segment",
    "backlog", "supply", "pricing", "price", "inventory", "customer", "capacity",
    "adoption", "usage", "consumption", "productivity", "migration", "cost savings", "roi",
    "competition", "competitive", "model", "inference", "ai", "utilization", "shipment",
)
_XBRL_CONCEPTS = (
    "revenue", "sales", "operatingincome", "operatingloss", "grossprofit", "profitloss", "netincome",
    "earningspershare", "eps", "cashflow", "cashandcash", "capitalexpenditure", "paymentsforproperty",
    "inventory", "accountsreceivable", "contractliabil", "remainingperformance", "backlog",
)
_FULL_COVERAGE_KINDS = {
    "transcript", "qa", "prepared_remarks", "financial_results", "financial_tables",
    "performance_review", "presentation", "supplement", "shareholder_letter",
}
_TRANSCRIPT_HINTS = ("transcript", "earnings call", "conference call", "prepared remarks", "q&a")
_FULL_TEXT_CAPS = {
    "transcript": 600_000,
    "qa": 600_000,
    "prepared_remarks": 600_000,
    "presentation": 320_000,
    "financial_results": 260_000,
    "financial_tables": 260_000,
    "performance_review": 320_000,
    "supplement": 320_000,
    "shareholder_letter": 320_000,
}


def _xbrl_facts(blob: bytes) -> list[dict]:
    facts: list[dict] = []
    try:
        with zipfile.ZipFile(io.BytesIO(blob)) as archive:
            names = [name for name in archive.namelist() if name.lower().endswith((".xbrl", ".xml"))]
            for name in names[:12]:
                try:
                    root = ET.fromstring(archive.read(name))
                except ET.ParseError:
                    continue
                for elem in root.iter():
                    context = elem.attrib.get("contextRef")
                    value = (elem.text or "").strip()
                    if not context or not value or len(value) > 100:
                        continue
                    tag = elem.tag.rsplit("}", 1)[-1]
                    if any(word in tag.casefold() for word in _XBRL_CONCEPTS):
                        facts.append({
                            "concept": tag, "value": value, "context": context,
                            "unit": elem.attrib.get("unitRef"), "scale": elem.attrib.get("scale"),
                            "source_file": PurePosixPath(name).name,
                        })
                        if len(facts) >= 200:
                            return facts
    except zipfile.BadZipFile:
        return []
    return facts


def _inline_xbrl_facts(blob: bytes) -> list[dict]:
    soup = BeautifulSoup(blob, "lxml")
    facts: list[dict] = []
    for tag in soup.find_all(True):
        if not tag.name.casefold().endswith("nonfraction"):
            continue
        concept = str(tag.get("name", ""))
        if not concept or not any(word in concept.casefold() for word in _XBRL_CONCEPTS):
            continue
        value = tag.get_text("", strip=True)
        if not value:
            continue
        facts.append({
            "concept": concept,
            "value": value,
            "context": tag.get("contextref") or tag.get("contextRef"),
            "unit": tag.get("unitref") or tag.get("unitRef"),
            "scale": tag.get("scale"),
            "source_file": "inline-xbrl",
        })
        if len(facts) >= 200:
            break
    return facts


def _xlsx_relevant_text(blob: bytes, max_chars: int = 60_000) -> str:
    workbook = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    selected: list[str] = []
    size = 0
    for sheet in workbook.worksheets[:30]:
        for row in sheet.iter_rows(max_row=1200, max_col=80, values_only=True):
            values = [str(value).strip() for value in row if value is not None and str(value).strip()]
            if not values:
                continue
            line = " | ".join(values)
            if any(keyword in line.casefold() for keyword in _KEYWORDS):
                value = f"[{sheet.title}] {line}"
                selected.append(value)
                size += len(value)
            if size >= max_chars:
                return "\n".join(selected)[:max_chars]
    return "\n".join(selected)[:max_chars]


def _effective_kind(disclosure: Disclosure) -> str:
    kind = str(disclosure.document_kind or "other").casefold()
    probe = f"{kind} {disclosure.title}".casefold()
    if any(hint in probe for hint in _TRANSCRIPT_HINTS):
        return "transcript" if kind not in {"qa", "prepared_remarks"} else kind
    return kind


def _bounded_full_text(text: str, cap: int) -> tuple[str, bool]:
    text = text or ""
    return text[:cap], len(text) > cap


def _pages_full_text(pages: list[str], cap: int) -> tuple[str, bool]:
    parts = [f"[[PAGE {index + 1}]]\n{text.strip()}" for index, text in enumerate(pages) if text.strip()]
    raw = "\n\n".join(parts)
    return _bounded_full_text(raw, cap)


def _metadata(
    disclosure: Disclosure,
    *,
    raw_chars: int,
    extracted_chars: int,
    truncated: bool,
    selection_mode: str,
    retrieval_excerpt: bool = False,
) -> dict:
    return {
        "source": disclosure.source,
        "document_kind": _effective_kind(disclosure),
        "raw_chars": raw_chars,
        "extracted_chars": extracted_chars,
        "truncated": truncated,
        "selection_mode": selection_mode,
        "retrieval_excerpt": retrieval_excerpt,
        "provenance": disclosure.metadata.get("provenance"),
        "qualitative_only": bool(disclosure.metadata.get("qualitative_only", False)),
    }


class EvidenceExtractor:
    def __init__(self, session: requests.Session | None = None):
        self.session = session if session is not None else requests.Session()

    def fetch(self, disclosure: Disclosure) -> Evidence:
        kind = _effective_kind(disclosure)

        if disclosure.source == "gemini_grounded_ir":
            text = str(disclosure.metadata.get("grounded_evidence", ""))
            return Evidence(
                disclosure.key, disclosure.title, disclosure.url, text,
                list(disclosure.metadata.get("structured_facts", []) or []),
                _metadata(
                    disclosure, raw_chars=len(text), extracted_chars=len(text), truncated=False,
                    selection_mode="retrieval_provider_excerpt", retrieval_excerpt=True,
                ),
            )

        if disclosure.source in {"alpha_vantage_transcript", "third_party_transcript"}:
            raw = str(disclosure.metadata.get("transcript_text", ""))
            selected, truncated = _bounded_full_text(raw, _FULL_TEXT_CAPS["transcript"])
            return Evidence(
                disclosure.key, disclosure.title, disclosure.url, selected, [],
                _metadata(
                    disclosure, raw_chars=len(raw), extracted_chars=len(selected), truncated=truncated,
                    selection_mode="full_transcript",
                ),
            )

        if not disclosure.document_url:
            return Evidence(
                disclosure.key, disclosure.title, disclosure.url, "", [],
                _metadata(disclosure, raw_chars=0, extracted_chars=0, truncated=False, selection_mode="no_document_url"),
            )

        user_agent = (
            os.getenv("SEC_USER_AGENT", "us-earnings-monitor/0.1 contact: research@example.com")
            if disclosure.source == "sec_edgar"
            else os.getenv("USER_AGENT", "Mozilla/5.0 earnings-monitor/0.5")
        )
        response = self.session.get(
            disclosure.document_url,
            timeout=45,
            headers={
                "User-Agent": user_agent,
                "Accept": "text/html,application/xhtml+xml,application/pdf;q=0.9,*/*;q=0.8",
                "Accept-Language": "en-US,en;q=0.9",
            },
        )
        response.raise_for_status()
        content_type = response.headers.get("content-type", "").casefold()
        blob = response.content
        bare_url = disclosure.document_url.casefold().split("?", 1)[0]

        if "spreadsheet" in content_type or bare_url.endswith(".xlsx"):
            selected = _xlsx_relevant_text(blob)
            return Evidence(
                disclosure.key, disclosure.title, disclosure.url, selected, [],
                _metadata(
                    disclosure, raw_chars=len(blob), extracted_chars=len(selected), truncated=False,
                    selection_mode="deterministic_relevant_spreadsheet_rows",
                ),
            )

        if "zip" in content_type or blob[:2] == b"PK":
            facts = _xbrl_facts(blob)
            return Evidence(
                disclosure.key, disclosure.title, disclosure.url, "", facts,
                _metadata(
                    disclosure, raw_chars=len(blob), extracted_chars=0, truncated=False,
                    selection_mode="structured_xbrl",
                ),
            )

        facts: list[dict] = []
        if "pdf" in content_type or bare_url.endswith(".pdf") or blob[:5] == b"%PDF-":
            reader = PdfReader(io.BytesIO(blob))
            pages = [(page.extract_text() or "") for page in reader.pages[:200]]
            raw_chars = sum(len(page) for page in pages)
            if kind in _FULL_COVERAGE_KINDS:
                selected_text, truncated = _pages_full_text(pages, _FULL_TEXT_CAPS.get(kind, 400_000))
                selection_mode = "full_document_pages"
            else:
                selected_text = compress_pdf_pages(pages, disclosure.document_kind, disclosure.title)
                truncated = False
                selection_mode = "deterministic_relevant_document_pages"
        else:
            text = BeautifulSoup(blob, "html.parser").get_text("\n", strip=True)
            raw_chars = len(text)
            facts = _inline_xbrl_facts(blob) if "html" in content_type or bare_url.endswith((".htm", ".html")) else []
            if kind in _FULL_COVERAGE_KINDS:
                selected_text, truncated = _bounded_full_text(text, _FULL_TEXT_CAPS.get(kind, 400_000))
                selection_mode = "full_document_text"
            else:
                selected_text = compress_text(text, disclosure.document_kind, disclosure.title)
                truncated = False
                selection_mode = "deterministic_relevant_document_text"

        return Evidence(
            disclosure.key,
            disclosure.title,
            disclosure.url,
            selected_text,
            facts,
            _metadata(
                disclosure,
                raw_chars=raw_chars,
                extracted_chars=len(selected_text),
                truncated=truncated,
                selection_mode=selection_mode,
            ),
        )
